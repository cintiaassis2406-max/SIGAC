from flask import send_file, request
from database.database import conectar

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

import tempfile
import re


def registrar_rotas(app):

    @app.route("/gerar_excel")
    def gerar_excel():

        conexao = conectar()
        cursor = conexao.cursor()

        credenciada_id = request.args.get("credenciada")
        mes = request.args.get("mes")
        ano = request.args.get("ano")
        tipo = request.args.get("tipo")

        situacao_financeira = request.args.get(
            "situacao_financeira",
            ""
        )

        nome_credenciada = ""

        if credenciada_id:

            cursor.execute(
                """
                SELECT nome
                FROM credenciadas
                WHERE id = %s
                """,
                (
                    credenciada_id,
                )
            )

            resultado_credenciada = cursor.fetchone()

            if resultado_credenciada:

                nome_credenciada = (
                    resultado_credenciada["nome"]
                )

        # ==================================================
        # ATENDIMENTOS
        # ==================================================

        sql = """
            SELECT
                a.colaborador,
                e.nome AS empresa,
                a.data_atendimento,
                a.tipo_atendimento,
                STRING_AGG(ae.nome_exame, ', ') AS exames,
                COALESCE(SUM(ae.valor_exame), 0) AS valor

            FROM atendimentos a

            INNER JOIN empresas e
                ON e.id = a.empresa_id

            LEFT JOIN atendimento_exames ae
                ON ae.atendimento_id = a.id

            WHERE 1=1
        """

        parametros = []

        # ==================================================
        # FILTRO CREDENCIADA
        # ==================================================

        if credenciada_id:

            sql += """
                AND a.credenciada_id = %s
            """

            parametros.append(
                credenciada_id
            )

        # ==================================================
        # FILTRO MÊS
        # ==================================================

        if mes:

            sql += """
                AND TO_CHAR(
                    a.data_atendimento,
                    'MM'
                ) = %s
            """

            parametros.append(
                f"{int(mes):02}"
            )

        # ==================================================
        # FILTRO ANO
        # ==================================================

        if ano:

            sql += """
                AND TO_CHAR(
                    a.data_atendimento,
                    'YYYY'
                ) = %s
            """

            parametros.append(
                str(ano)
            )

        # ==================================================
        # AGRUPAMENTO
        # ==================================================

        sql += """
            GROUP BY
                a.id,
                e.nome

            ORDER BY
                a.colaborador
        """

        cursor.execute(
            sql,
            parametros
        )

        dados = cursor.fetchall()

        # ==================================================
        # DETALHAMENTO DOS EXAMES
        # ==================================================

        sql_exames = """
            SELECT
                ae.nome_exame,
                ae.valor_exame,
                COUNT(*) AS quantidade,
                SUM(ae.valor_exame) AS total

            FROM atendimento_exames ae

            INNER JOIN atendimentos a
                ON a.id = ae.atendimento_id

            WHERE 1=1
        """

        parametros_exames = []

        # ==================================================
        # FILTRO CREDENCIADA
        # ==================================================

        if credenciada_id:

            sql_exames += """
                AND a.credenciada_id = %s
            """

            parametros_exames.append(
                credenciada_id
            )

        # ==================================================
        # FILTRO MÊS
        # ==================================================

        if mes:

            sql_exames += """
                AND TO_CHAR(
                    a.data_atendimento,
                    'MM'
                ) = %s
            """

            parametros_exames.append(
                f"{int(mes):02}"
            )

        # ==================================================
        # FILTRO ANO
        # ==================================================

        if ano:

            sql_exames += """
                AND TO_CHAR(
                    a.data_atendimento,
                    'YYYY'
                ) = %s
            """

            parametros_exames.append(
                str(ano)
            )

        # ==================================================
        # FILTRO TIPO DE ATENDIMENTO
        # ==================================================

        if tipo:

            sql_exames += """
                AND a.tipo_atendimento = %s
            """

            parametros_exames.append(
                tipo
            )
        # ==================================================
        # FILTRO SITUAÇÃO FINANCEIRA
        # ==================================================

        # Não aplicar este filtro no detalhamento dos exames.
        # O detalhamento deve acompanhar os atendimentos
        # selecionados pelo período e pela credenciada.

        # ==================================================
        # AGRUPAMENTO DOS EXAMES
        # ==================================================

        sql_exames += """
            GROUP BY
                ae.nome_exame,
                ae.valor_exame

            ORDER BY
                ae.nome_exame
        """

        cursor.execute(
            sql_exames,
            parametros_exames
        )

        detalhamento_exames = cursor.fetchall()

        # ==================================================
        # CRIAR EXCEL
        # ==================================================

        wb = Workbook()

        ws = wb.active

        ws.title = "Faturamento"

        # ==================================================
        # CABEÇALHO
        # ==================================================

        ws.append([
            "SIGAC - RELATÓRIO DE FATURAMENTO"
        ])

        ws.append([
            f"Credenciada: {nome_credenciada}"
        ])

        ws.append([
            f"Período: {mes or ''}/{ano or ''}"
        ])

        ws.append([
            f"Situação: {situacao_financeira or 'Todas'}"
        ])

        ws.append([])

        # ==================================================
        # TABELA DE ATENDIMENTOS
        # ==================================================

        ws.append([
            "Colaborador",
            "Empresa",
            "Data",
            "Tipo Atendimento",
            "Exames",
            "Valor"
        ])

        # ==================================================
        # FORMATAÇÃO DO CABEÇALHO
        # ==================================================

        for celula in ws[6]:

            celula.font = Font(
                bold=True
            )

            celula.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        # ==================================================
        # DADOS DOS ATENDIMENTOS
        # ==================================================

        total_geral = 0

        for item in dados:

            if item["data_atendimento"]:

                if hasattr(
                    item["data_atendimento"],
                    "strftime"
                ):

                    data_atendimento = (
                        item["data_atendimento"]
                        .strftime("%d/%m/%Y")
                    )

                else:

                    data_atendimento = str(
                        item["data_atendimento"]
                    )

            else:

                data_atendimento = ""

            valor = item["valor"] or 0

            ws.append([
                item["colaborador"] or "",
                item["empresa"] or "",
                data_atendimento,
                item["tipo_atendimento"] or "",
                item["exames"] or "",
                float(valor)
            ])

            total_geral += float(valor)

        # ==================================================
        # TOTAL DOS ATENDIMENTOS
        # ==================================================

        ws.append([])

        ws.append([
            "",
            "",
            "",
            "",
            "TOTAL",
            total_geral
        ])

        ultima_linha_atendimentos = ws.max_row

        # ==================================================
        # FORMATAÇÃO DOS DADOS
        # ==================================================

        for linha in ws.iter_rows():

            for celula in linha:

                celula.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

        # ==================================================
        # FORMATAÇÃO DO TOTAL
        # ==================================================

        ws.cell(
            row=ultima_linha_atendimentos,
            column=5
        ).font = Font(
            bold=True
        )

        ws.cell(
            row=ultima_linha_atendimentos,
            column=6
        ).font = Font(
            bold=True
        )

        # ==================================================
        # FORMATO DOS VALORES
        # ==================================================

        for linha in range(
            7,
            ultima_linha_atendimentos + 1
        ):

            ws.cell(
                row=linha,
                column=6
            ).number_format = (
                'R$ #,##0.00'
            )

        # ==================================================
        # DETALHAMENTO DOS EXAMES
        # ==================================================

        ws.append([])

        ws.append([
            "DETALHAMENTO DOS EXAMES REALIZADOS"
        ])

        ws.append([])

        ws.append([
            "Exame",
            "Valor Unitário",
            "Quantidade",
            "Total"
        ])

        linha_cabecalho_exames = ws.max_row

        # ==================================================
        # FORMATAÇÃO DO CABEÇALHO DOS EXAMES
        # ==================================================

        for celula in ws[linha_cabecalho_exames]:

            celula.font = Font(
                bold=True
            )

            celula.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        # ==================================================
        # DADOS DOS EXAMES
        # ==================================================

        total_exames = 0

        for exame in detalhamento_exames:

            nome_exame = (
                exame["nome_exame"]
                or ""
            )

            valor_unitario = (
                exame["valor_exame"]
                or 0
            )

            quantidade = (
                exame["quantidade"]
                or 0
            )

            total = (
                exame["total"]
                or 0
            )

            ws.append([
                nome_exame,
                float(valor_unitario),
                quantidade,
                float(total)
            ])

            total_exames += float(total)

        # ==================================================
        # TOTAL DOS EXAMES
        # ==================================================

        ws.append([])

        ws.append([
            "",
            "",
            "TOTAL",
            total_exames
        ])

        ultima_linha = ws.max_row

        # ==================================================
        # FORMATAÇÃO DOS VALORES DOS EXAMES
        # ==================================================

        for linha in range(
            linha_cabecalho_exames + 1,
            ultima_linha + 1
        ):

            ws.cell(
                row=linha,
                column=2
            ).number_format = (
                'R$ #,##0.00'
            )

            ws.cell(
                row=linha,
                column=4
            ).number_format = (
                'R$ #,##0.00'
            )

        # ==================================================
        # FORMATAÇÃO DO TOTAL DOS EXAMES
        # ==================================================

        ws.cell(
            row=ultima_linha,
            column=3
        ).font = Font(
            bold=True
        )

        ws.cell(
            row=ultima_linha,
            column=4
        ).font = Font(
            bold=True
        )

        # ==================================================
        # LARGURA DAS COLUNAS
        # ==================================================

        larguras = {
            "A": 32,
            "B": 35,
            "C": 18,
            "D": 20,
            "E": 60,
            "F": 18
        }

        for coluna, largura in larguras.items():

            ws.column_dimensions[
                coluna
            ].width = largura

        # ==================================================
        # ALTURA DAS LINHAS
        # ==================================================

        for linha in range(
            7,
            ws.max_row + 1
        ):

            ws.row_dimensions[
                linha
            ].height = 30

        # ==================================================
        # CONGELAR CABEÇALHO
        # ==================================================

        ws.freeze_panes = "A7"

        # ==================================================
        # FILTRO NO CABEÇALHO
        # ==================================================

        if ultima_linha_atendimentos >= 6:

            ws.auto_filter.ref = (
                f"A6:F{ultima_linha_atendimentos}"
            )

        # ==================================================
        # CRIAR ARQUIVO
        # ==================================================

        arquivo = tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx"
        )

        wb.save(
            arquivo.name
        )

        conexao.close()

        nome_arquivo = (
            f"{nome_credenciada} - "
            f"{mes or 'Todos'}-{ano or ''} - "
            f"{tipo or 'Faturar'}.xlsx"
        )

        nome_arquivo = re.sub(
            r'[\\/:\*?"<>|]',
            '',
            nome_arquivo
        )

        return send_file(
            arquivo.name,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )